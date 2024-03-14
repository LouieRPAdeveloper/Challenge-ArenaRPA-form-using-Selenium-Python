from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait

from selenium.webdriver.support import expected_conditions as EC
import time

from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

from selenium.webdriver.common.keys import Keys

import openpyxl

import yagmail

#importamos todas las librerias correspondientes

#ingresamos el path, previamente tenemos que descargar el archivo
path ="C:/Users/alber/Downloads/ArenaRPA_FormData.xlsx"

#ingresamos al workbook
workbook=openpyxl.load_workbook(path)

#se le asigna sheet a  asigna la hoja activa del libro de trabajo 
sheet=workbook.active
 

rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)
print("pau")

#creamos listas vacias en referencia a la cantidad de columnas y al nombre de cada 
#uno de ellas

Nombres_excel=[]
Apellidos_excel=[]
Empresa_excel=[]
Numero_excel=[]
Email_excel=[]
Pais_excel=[]
Web_excel=[]


#recorremos primero las filas y columnas

for r in range (2, rows+1):
    
    for c in range(1,cols+1):
        
        #lo que cambia es las columnas, por lo que se le asigna el valor de 1 al nombre
        if (c == 1):
            Nombres_excel.append(sheet.cell(row=r,column=c).value)
        # el valor de 2 se le asigna al apellido
        elif (c == 2):
            Apellidos_excel.append(sheet.cell(row=r,column=c).value)
        
        #el valor de 3 a la empresa
        elif (c == 3):
            Empresa_excel.append(sheet.cell(row=r,column=c).value)
        #el valor de 4 al numero telefonico
        elif (c == 4):
            Numero_excel.append(sheet.cell(row=r,column=c).value)
        #el valor de 5 al email
        elif (c == 5):
            Email_excel.append(sheet.cell(row=r,column=c).value)
        #el valor de 6 al pais
        elif (c == 6):
            Pais_excel.append(sheet.cell(row=r,column=c).value)
        #el valor de 7 a la web
        elif (c == 7):
            Web_excel.append(sheet.cell(row=r,column=c).value)


#vemos como ejemplos la lista completa de los nombres, apellidos, web
print(Nombres_excel)
print(len(Nombres_excel))

print(Apellidos_excel)
print(len(Apellidos_excel))

print(Web_excel)
print(len(Web_excel))


# Creamos el objeto WebDriver
driver = webdriver.Chrome()

# Navegamos a la página de web del reto
driver.get("https://arenarpa.com/crazy-form")

driver.maximize_window()

#espera que cargue la pagina por 5 segundos
time.sleep(5)

#boton que da clic a iniciar reto
#nos guiamos del XPath
boton_INICIAR_RETO = driver.find_element(By.XPATH,"/html/body/app-root/app-crazy-form/div/div[1]/div[2]/div[2]/a")
boton_INICIAR_RETO.click()

#recorremos desde 0 hasta el tamano de las listas
#ya que range(0,len) toma de 0 a len -1
for i in range (0, len(Nombres_excel)):
    
    #buscamos el elemento por el ID nombres
    input_nombres = driver.find_element(By.ID,"nombres")
    #Asignamos el valor del nombre
    input_nombres.send_keys(Nombres_excel[i])
    
    #buscamos el elemento por el ID apellidos
    input_apellidos = driver.find_element(By.ID,"apellidos")
    #Asignamos el valor del apellido
    input_apellidos.send_keys(Apellidos_excel[i])
    
    #buscamos el elemento por el ID empresa
    input_empresa=driver.find_element(By.ID,"empresa")
    #Asignamos el valor de empresa
    input_empresa.send_keys(Empresa_excel[i])
    
    #buscamos el elemento por el ID numero
    input_numero=driver.find_element(By.ID,"numero")
    #Asignamos el valor del numero telefonico
    input_numero.send_keys(Numero_excel[i])
    
    #buscamos el elemento por el ID email
    input_email = driver.find_element(By.ID,"email")
    #Asignamos el valor del email
    input_email.send_keys(Email_excel[i])
    
    #buscamos el elemento por el ID pais
    input_pais=driver.find_element(By.ID,"pais")
    #Asignamos el valor del pais
    input_pais.send_keys(Pais_excel[i])
    
    #buscamos el elemento por el ID web
    input_web=driver.find_element(By.ID,"web")
    #Asignamos el valor de la web
    input_web.send_keys(Web_excel[i])
    
    
    #buscamos el elemento mediante el CSS_SELECTOR debido a que
    #previamente se tuvo que analizar el boton, ya que el boton 
    #enviar cambiaba constantemente
    boton_envio_reto1 = driver.find_element(By.CSS_SELECTOR, 'button.text-white')
    
    #da clic al encontrar el boton
    boton_envio_reto1.click()
    
   
    #hacemos que tome un tiempo de 2 segundos
    time.sleep(2)

#hacemos una espera de 10 segundos
time.sleep(5)
#guarda el pantallazo en la carpeta indicada
driver.save_screenshot("D:/Python/foto1.png")
